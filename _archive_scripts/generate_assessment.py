"""Generate the filled comparative analysis Excel file."""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.load_workbook("C:/Users/HP/Downloads/PROC AI FNF COMPARATIVE ANALYSIS.xlsx")
ws = wb.active

# Answers: row_number -> (scope, explanation)
answers = {
    3: ("YES", "Built. ChatPage + OrchestratorAgent classifies intent via NL chat. BudgetVerificationAgent checks budget. VendorSelectionAgent creates PR. Full SSE streaming. Gap: Spend classification is keyword-based, could enhance with ML for UNSPSC codes."),
    4: ("PARTIAL", "SpendAnalyticsAgent identifies vendor concentration and duplicates. Missing: PR-level de-duplication across departments. Need: consolidation agent scanning open PRs for overlapping items. Effort: 2-3 days."),
    5: ("PARTIAL", "BudgetVerificationAgent checks budget availability (point-in-time). budget_tracking table has committed vs available. DashboardPage shows utilization. Missing: ML-based forecasting to predict end-of-period overrun. Need: time-series model. Effort: 3-5 days."),
    6: ("YES", "Built. ApprovalRoutingAgent + approval_rules table with configurable thresholds by document_type and amount range. Auto-escalation via sla_hours and escalate_after. PaymentApprovalAgent handles payment-level rules separately. Minor gap: absent-approver delegation. Effort: 1 day."),
    7: ("PARTIAL", "VendorSelectionAgent scores vendors (quality 40%, price 30%, delivery 20%, category 10%). SanctionsService checks blocklists. Missing: external crawling for ESG, news, trade registries. Need: external APIs (Dun & Bradstreet, OpenCorporates). Effort: 5-7 days."),
    8: ("PARTIAL", "QuoteComparisonAgent scores bids with weighted criteria. rfq_headers, vendor_quotes, quote_comparisons tables exist. Missing: auto-generation of RFP/RFQ documents. Need: LLM-based template generation. Effort: 3-4 days."),
    9: ("NO", "PriceAnalysisAgent does price benchmarking. Historical contract data exists. No real-time negotiation copilot. Need: LLM-based advisory agent with market data + historical contracts + supplier financials. Effort: 5-7 days."),
    10: ("NO", "Product tables exist. PriceAnalysisAgent compares prices. No continuous catalog monitoring. Need: scheduled agent monitoring price changes, discontinuations, policy violations. Effort: 3-4 days."),
    11: ("PARTIAL", "OrchestratorAgent classifies query intent. SpendAnalyticsAgent categorizes by vendor/department. Missing: UNSPSC 6-digit classification at >95% accuracy. Need: fine-tuned ML classifier or LLM prompt. Effort: 3-5 days."),
    12: ("YES", "Built. VendorSelectionAgent converts PRs to POs with supplier selection and contract pricing. POIntakeAgent + PORegistrationAgent validate. Odoo PO creation via XML-RPC. >70% routine orders automated. Gap: multi-delivery-location splitting. Effort: 1-2 days."),
    13: ("PARTIAL", "po_amendments table exists. BudgetVerificationAgent checks budget. Missing: agent modeling cascading downstream impact of PO changes on GRN/invoice/budget. Need: impact analysis agent. Effort: 3-4 days."),
    14: ("PARTIAL", "NotificationDispatcher + 19 email_templates + notification_log built. EmailService supports SendGrid/SMTP/Mock. Missing: auto-chasing for delivery confirmations. Notifications in mock mode. Need: activate sending + chasing scheduler. Effort: 2-3 days."),
    15: ("PARTIAL", "CycleTimeReportPage in frontend. po_approval_log, pr_approval_workflows tables track timing. Missing: backend API for cycle-time analytics with AI bottleneck recommendations. Effort: 1-2 days."),
    16: ("YES", "Built. VendorOnboardingPage multi-step wizard. SanctionsService for compliance screening. vendor_evaluations table. Gaps: document upload local-only (no cloud storage), tax/insurance validation basic. Effort to enhance: 2-3 days."),
    17: ("PARTIAL", "RiskAssessmentAgent with 4-dimension scoring. SanctionsService (Local/OpenSanctions/WorldBank). RiskAssessmentPage + po_risk_assessments (295 records). Missing: external real-time signals (news, ESG, geopolitical, cyber). Need: external APIs. Effort: 5-7 days."),
    18: ("PARTIAL", "SupplierPerformanceAgent + SupplierPerformancePage with quadrant analysis. vendor_performance tables. Missing: auto-generated improvement plans sent to suppliers. Need: LLM-generated report + email dispatch. Effort: 2-3 days."),
    19: ("NO", "Backend APIs exist for all data. No supplier-facing portal. Need: separate frontend app with supplier auth, PO viewing, invoice submission, payment tracking, AI chatbot. Effort: 2-4 weeks."),
    20: ("YES", "Built. InvoiceCaptureAgent with pluggable OCR (Regex/Mindee/AWS Textract). PDF/email/portal support. ocr_ingestion_log table. DocumentProcessingPage. Gap: EDI not built. Email inbox agent exists but unwired (Sprint 9). Effort: 2-3 days."),
    21: ("YES", "Built. InvoiceMatchingAgent with configurable thresholds (<=5% auto, 5-10% flag, >20% block). DiscrepancyResolutionAgent auto-resolves. three_way_match_log, discrepancy_log tables. Gap: line-level matching basic. Effort: 2-3 days."),
    22: ("PARTIAL", "GoodsReceiptPage 3-step wizard. GoodsReceiptAgent handles full/partial/quality. grn_headers, grn_line_items, qc_inspection_log tables. Missing: mobile app, barcode/QR, photo capture. Need: React Native or PWA. Effort: 2-3 weeks."),
    23: ("YES", "Built. DiscrepancyResolutionAgent classifies by root cause (price/qty mismatch, missing PO, duplicate). Auto-resolves tolerables. Routes rest to manual queue. invoice_exceptions, discrepancy_log tables. Enhancement: LLM resolution suggestions. Effort: 1-2 days."),
    24: ("PARTIAL", "InvoiceCaptureAgent checks exact duplicates. duplicate_invoice_log table. AnomalyDetectionPage. AnomalyDetectionAgent exists (needs wiring). Missing: fuzzy matching, amount inflation, fictitious supplier detection. Effort: 3-4 days."),
    25: ("YES", "Built. PaymentCalculationAgent applies early-payment discounts from contracts. FX conversion via pluggable FXService. early_payment_discounts table. Gap: DPO target optimization, cash flow forecasting. Effort: 2-3 days."),
    26: ("PARTIAL", "PaymentExecutionService with manual/bank_api/ach modes. payment_runs/payment_run_lines tables. PaymentExecutionPage. Gap: bank API and ACH are stubs, only manual works. Need: bank integration. Effort: 5-10 days."),
    27: ("NO", "AgingService generates AP aging reports. ap_aging table. AgingReportPage. Missing: bank statement import (MT940/BAI2), auto-reconciliation matching, unmatched investigation. Effort: 5-7 days."),
    28: ("NO", "Early payment discount calculation exists. No marketplace UI for supplier-initiated discount requests. Need: supplier-facing interface + buyer approval. Effort: 2-3 weeks."),
    29: ("PARTIAL", "DiscrepancyResolutionAgent + InvoiceRoutingAgent handle exceptions per-module. discrepancy_log, invoice_holds tables. Not unified into single inbox. Need: universal exception queue + AI triage. Effort: 3-4 days."),
    30: ("PARTIAL", "InvoiceMatchingAgent has static thresholds (5%/10%/20%). Not configurable per category/supplier. Need: tolerance config table + ML optimization. Effort: 3-5 days."),
    31: ("YES", "Built. ComplianceCheckAgent validates policy. PaymentReadinessAgent has 7-gate check. approval_rules configurable. Gap: rules not editable via UI. Need: policy editor page. Effort: 2-3 days."),
    32: ("PARTIAL", "audit_trail table + agent_actions (5,507 records). AnomalyDetectionAgent exists. Missing: SOD conflict detection, continuous monitoring loop, auto-audit reports. Need: SOD rules engine. Effort: 4-5 days."),
    33: ("YES", "Built. TaxService: UAE VAT 5%, Saudi VAT 15%, EU VAT, US sales tax, India GST, Singapore GST, Japan CT. Zero-rated/exempt handling. tax_codes table. Gap: e-Invoicing mandates (ZATCA, India GST). Effort: 5-7 days per country."),
    34: ("NO", "SanctionsService has basic blocklist only. No ESG data model, carbon tracking, diversity metrics, sustainability reporting. Need: ESG tables + scoring agent + dashboard. Effort: 5-7 days."),
    35: ("YES", "Built. SpendAnalyticsAgent with savings identification (consolidation, renegotiation, price variance). SpendAnalyticsPage with charts. spend_analytics, budget_vs_actuals tables. Gap: historical savings tracking. Effort: 1-2 days."),
    36: ("PARTIAL", "FXService with 18 currencies (Static/DB/OpenExchangeRates). exchange_rates table. Missing: commodity prices, energy costs, supplier inflation indices. Need: commodity API (Quandl/Trading Economics). Effort: 3-5 days."),
    37: ("NO", "SystemHealthPage shows current health. Missing: predictive forecasting on process metrics. Need: time-series agent. Effort: 3-5 days."),
    38: ("YES", "Core feature. ChatPage with NL chat. QueryRouter + LLM classification. ConversationalHandler generates responses. SSE streaming. Multi-language (Arabic, Urdu). SQL generation for ad-hoc queries. Strongest feature in the platform."),
    39: ("NO", "agent_actions has 5,507 event records (rich log). audit_trail exists. Missing: process mining engine mapping actual flows, identifying deviations, quantifying rework. Need: event log analysis agent + visualization. Effort: 5-7 days."),
    40: ("NO", "contracts table + ContractMonitoringAgent tracks expiry. ContractMonitoringPage. Missing: AI drafting from templates, clause analysis, non-standard clause flagging. Need: LLM contract generation agent. Effort: 5-7 days."),
    41: ("PARTIAL", "ContractMonitoringAgent monitors expiry/renewal dates. contracts table with terms. ContractMonitoringPage. Missing: NLP extraction of obligations/SLAs from contract PDFs. Need: document parsing pipeline. Effort: 5-7 days."),
    42: ("YES", "Built. OrchestratorAgent master router for 21+ agents. PipelineOrchestrator runs 9-step I2P pipeline with context chaining. ODAL pattern. SSE agent event streaming. PipelineVisualizerPage. Enhancement: LangGraph for state machines. Effort: 3-5 days."),
    43: ("PARTIAL", "Every agent returns confidence_score (0.0-1.0). agent_actions logs decisions. agent_decisions has reasoning field. PendingApprovalsPage shows reasoning. Missing: NL narrative explanations. Need: LLM post-processing. Effort: 2-3 days."),
    44: ("NO", "Agent actions logged (5,507). Approval decisions recorded. No feedback loop, no model retraining, no outcome tracking. Need: feedback collection + fine-tuning pipeline. Effort: 1-2 weeks."),
    45: ("NO", "No in-app guidance or adoption analytics. Need: product tour (Shepherd.js/Intro.js) + contextual tooltips + usage tracking. Effort: 3-5 days."),
}

# Styles
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font = Font(color="006100", bold=True)
yellow_font = Font(color="9C6500", bold=True)
red_font = Font(color="9C0006", bold=True)
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# Headers
ws.cell(row=1, column=10, value="Proc AI Scope").font = Font(bold=True, size=11)
ws.cell(row=1, column=10).alignment = wrap
ws.cell(row=1, column=10).border = border
ws.cell(row=1, column=11, value="Assessment: What Exists / What's Missing / Effort").font = Font(bold=True, size=11)
ws.cell(row=1, column=11).alignment = wrap
ws.cell(row=1, column=11).border = border

ws.column_dimensions['J'].width = 16
ws.column_dimensions['K'].width = 85

# Fill answers
for row_num, (scope, detail) in answers.items():
    cj = ws.cell(row=row_num, column=10, value=scope)
    cj.alignment = Alignment(horizontal="center", vertical="top")
    cj.border = border
    if scope == "YES":
        cj.fill, cj.font = green_fill, green_font
    elif scope == "PARTIAL":
        cj.fill, cj.font = yellow_fill, yellow_font
    else:
        cj.fill, cj.font = red_fill, red_font

    ck = ws.cell(row=row_num, column=11, value=detail)
    ck.alignment = wrap
    ck.border = border
    ck.font = Font(size=10)

# Summary
sr = max(answers.keys()) + 2
ws.cell(row=sr, column=1, value="SUMMARY").font = Font(bold=True, size=13)

for i, (label, val, fill, fnt) in enumerate([
    ("Total Features", "43", None, Font(bold=True)),
    ("YES (Fully Built)", "13 (30%)", green_fill, green_font),
    ("PARTIAL (Needs Enhancement)", "19 (44%)", yellow_fill, yellow_font),
    ("NO (Not Built)", "11 (26%)", red_fill, red_font),
    ("Coverage (YES + PARTIAL)", "74%", None, Font(bold=True, size=13)),
]):
    ws.cell(row=sr+1+i, column=1, value=label).font = Font(bold=True)
    c = ws.cell(row=sr+1+i, column=10, value=val)
    c.font = fnt
    if fill:
        c.fill = fill

out = "C:/Users/HP/Downloads/PROC_AI_FNF_COMPARATIVE_ANALYSIS_FILLED.xlsx"
wb.save(out)
print(f"Saved: {out}")

yes = sum(1 for s, _ in answers.values() if s == "YES")
partial = sum(1 for s, _ in answers.values() if s == "PARTIAL")
no = sum(1 for s, _ in answers.values() if s == "NO")
print(f"YES: {yes} | PARTIAL: {partial} | NO: {no} | Total: {len(answers)}")
