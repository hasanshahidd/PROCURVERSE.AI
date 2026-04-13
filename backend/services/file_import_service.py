"""
File Import Service
====================
Handles Excel (.xlsx/.xls) and CSV file uploads:
  1. Reads file headers + data
  2. Auto-generates PostgreSQL table schema (or adds columns to existing table)
  3. Imports all rows
  4. Returns import summary

Used by:
  - POST /api/import/upload        — single file upload
  - POST /api/import/upload-batch  — multiple file upload
  - POST /api/import/preview       — preview schema without importing
"""

import os
import csv
import io
import re
import logging
from pathlib import Path
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

import psycopg2
from psycopg2.extras import execute_values

log = logging.getLogger(__name__)


# ── Type Inference ───────────────────────────────────────────────────────────

DIRTY_PREFIXES = ('INVALID', 'TEMP', 'ERROR', 'XXX', 'TBD', 'UNKNOWN')

def infer_pg_type(values: list[str]) -> str:
    """Infer PostgreSQL column type from values. Uses TEXT for any dirty/mixed data."""
    all_vals = [v.strip() for v in values if v.strip()]
    clean = [v for v in all_vals if v.upper() not in ('NULL', 'N/A', '', 'NONE', 'TBD', 'UNKNOWN')]
    if not clean:
        return 'TEXT'

    has_dirty = any(v.upper().startswith(DIRTY_PREFIXES) for v in all_vals)

    # Integer check
    if not has_dirty:
        try:
            ints = [int(v) for v in clean]
            max_val = max(abs(x) for x in ints)
            return 'BIGINT' if max_val > 2_147_483_647 else 'INTEGER'
        except ValueError:
            pass

    # Decimal check
    if not has_dirty:
        try:
            [Decimal(v.replace(',', '')) for v in clean]
            return 'NUMERIC(18,4)'
        except (InvalidOperation, ValueError):
            pass

    # Date check
    if not has_dirty and len(clean) > 2:
        date_patterns = [r'^\d{4}-\d{2}-\d{2}', r'^\d{2}/\d{2}/\d{4}']
        date_count = sum(1 for v in clean if any(re.match(p, v) for p in date_patterns))
        if date_count >= len(clean) * 0.9:
            if any('T' in v or ' ' in v.strip() for v in clean[:10]):
                return 'TIMESTAMP'
            return 'DATE'

    return 'TEXT'


def sanitize_column_name(col: str) -> str:
    """Convert any header string to a valid PostgreSQL column name."""
    col = col.strip().strip('\ufeff')
    col = re.sub(r'[^a-zA-Z0-9_]', '_', col)
    col = re.sub(r'_+', '_', col).strip('_').lower()
    if col and col[0].isdigit():
        col = 'c_' + col
    reserved = {'order', 'group', 'user', 'table', 'select', 'where', 'from',
                'index', 'key', 'check', 'default', 'column', 'row', 'level',
                'type', 'comment', 'primary', 'references', 'constraint'}
    if col in reserved:
        col = col + '_val'
    return col or 'unnamed_col'


def sanitize_table_name(name: str) -> str:
    """Convert filename or user-provided name to valid PostgreSQL table name."""
    name = Path(name).stem.lower()
    name = re.sub(r'[^a-z0-9_]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    if not name or name[0].isdigit():
        name = 'imported_' + name
    return name


# ── File Readers ─────────────────────────────────────────────────────────────

def read_csv_bytes(file_bytes: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Read CSV from bytes, return (headers, rows)."""
    text = file_bytes.decode('utf-8-sig')
    reader = csv.reader(io.StringIO(text))
    headers = next(reader)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return headers, rows


def read_excel_bytes(file_bytes: bytes, filename: str, sheet_name: str = None) -> tuple[list[str], list[list[str]]]:
    """Read Excel from bytes, return (headers, rows). Requires openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return [], []

    headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(rows_raw[0])]
    rows = []
    for row in rows_raw[1:]:
        str_row = [str(cell) if cell is not None else '' for cell in row]
        if any(cell.strip() for cell in str_row):
            rows.append(str_row)

    return headers, rows


def read_file(file_bytes: bytes, filename: str, sheet_name: str = None) -> tuple[list[str], list[list[str]]]:
    """Auto-detect file type and read headers + rows."""
    ext = Path(filename).suffix.lower()
    if ext in ('.xlsx', '.xls', '.xlsm'):
        return read_excel_bytes(file_bytes, filename, sheet_name)
    elif ext in ('.csv', '.tsv'):
        return read_csv_bytes(file_bytes, filename)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .csv, .xlsx, .xls")


# ── Schema Generation ────────────────────────────────────────────────────────

def generate_schema(table_name: str, headers: list[str], rows: list[list[str]]) -> dict:
    """Generate table schema from CSV/Excel headers + sample data.

    Returns dict with:
      - table_name: sanitized name
      - columns: list of {name, original_header, pg_type}
      - ddl: CREATE TABLE SQL
      - row_count: number of data rows
    """
    columns = []
    for i, raw_header in enumerate(headers):
        col_name = sanitize_column_name(raw_header)
        sample_values = [row[i] for row in rows if i < len(row)]
        pg_type = infer_pg_type(sample_values)
        columns.append({
            'name': col_name,
            'original_header': raw_header.strip(),
            'pg_type': pg_type,
        })

    col_defs = ',\n'.join(f'    {c["name"]} {c["pg_type"]}' for c in columns)
    ddl = f'CREATE TABLE IF NOT EXISTS {table_name} (\n    _row_id SERIAL PRIMARY KEY,\n{col_defs}\n);'

    return {
        'table_name': table_name,
        'columns': columns,
        'ddl': ddl,
        'row_count': len(rows),
    }


# ── Import Engine ────────────────────────────────────────────────────────────

def import_data(
    db_url: str,
    file_bytes: bytes,
    filename: str,
    table_name: Optional[str] = None,
    sheet_name: Optional[str] = None,
    mode: str = 'replace',  # 'replace' | 'append' | 'skip_existing'
    dry_run: bool = False,
) -> dict:
    """Import a file into PostgreSQL.

    Args:
        db_url: PostgreSQL connection string
        file_bytes: raw file content
        filename: original filename (for type detection)
        table_name: override table name (default: derived from filename)
        sheet_name: Excel sheet name (default: active sheet)
        mode: 'replace' (DROP+CREATE), 'append' (INSERT into existing), 'skip_existing' (skip if table exists)
        dry_run: if True, return schema without importing

    Returns:
        dict with import results
    """
    # Read file
    headers, rows = read_file(file_bytes, filename, sheet_name)
    if not headers:
        return {'success': False, 'error': 'File has no headers/data', 'rows_imported': 0}

    # Determine table name
    tbl = sanitize_table_name(table_name or filename)

    # Generate schema
    schema = generate_schema(tbl, headers, rows)

    if dry_run:
        return {
            'success': True,
            'dry_run': True,
            'table_name': tbl,
            'schema': schema,
            'rows_to_import': len(rows),
        }

    # Connect and import
    conn = psycopg2.connect(db_url)
    cur = conn.cursor()
    col_names = [c['name'] for c in schema['columns']]

    try:
        # Check if table exists
        cur.execute("SELECT EXISTS(SELECT 1 FROM information_schema.tables WHERE table_schema='public' AND table_name=%s)", (tbl,))
        table_exists = cur.fetchone()[0]

        if mode == 'skip_existing' and table_exists:
            return {
                'success': True,
                'skipped': True,
                'table_name': tbl,
                'message': f'Table {tbl} already exists, skipped',
                'rows_imported': 0,
            }

        if mode == 'replace':
            cur.execute(f'DROP TABLE IF EXISTS {tbl} CASCADE;')
            cur.execute(schema['ddl'])
        elif mode == 'append' and not table_exists:
            cur.execute(schema['ddl'])
        # If append and table exists, just insert

        # Import rows
        if rows:
            cols_joined = ', '.join(col_names)
            template = '(' + ', '.join(['%s'] * len(col_names)) + ')'
            cleaned_rows = []
            for row in rows:
                cleaned = []
                for i, val in enumerate(row):
                    if i >= len(col_names):
                        break
                    val = val.strip() if val else ''
                    if val == '' or val.upper() in ('NULL', 'NONE'):
                        cleaned.append(None)
                    else:
                        cleaned.append(val)
                while len(cleaned) < len(col_names):
                    cleaned.append(None)
                cleaned_rows.append(tuple(cleaned))

            execute_values(
                cur,
                f'INSERT INTO {tbl} ({cols_joined}) VALUES %s',
                cleaned_rows,
                template=template,
                page_size=500
            )

        conn.commit()

        return {
            'success': True,
            'table_name': tbl,
            'columns': len(col_names),
            'rows_imported': len(rows),
            'mode': mode,
            'schema': schema,
        }

    except Exception as e:
        conn.rollback()
        log.error(f"Import error for {tbl}: {e}")
        return {
            'success': False,
            'table_name': tbl,
            'error': str(e),
            'rows_imported': 0,
        }
    finally:
        cur.close()
        conn.close()


def import_batch(
    db_url: str,
    files: list[tuple[bytes, str]],  # list of (file_bytes, filename)
    mode: str = 'replace',
    dry_run: bool = False,
) -> dict:
    """Import multiple files at once. Returns summary."""
    results = []
    total_rows = 0
    total_tables = 0
    errors = []

    for file_bytes, filename in files:
        result = import_data(db_url, file_bytes, filename, mode=mode, dry_run=dry_run)
        results.append(result)
        if result.get('success'):
            total_rows += result.get('rows_imported', 0)
            total_tables += 1
        else:
            errors.append({'filename': filename, 'error': result.get('error', 'unknown')})

    return {
        'success': len(errors) == 0,
        'total_files': len(files),
        'total_tables': total_tables,
        'total_rows': total_rows,
        'errors': errors,
        'results': results,
    }
